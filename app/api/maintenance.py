"""Maintenance API — payments, budgets, service fees, rental schedule import."""

import io
import re
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from pydantic import BaseModel
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_user
from app.database import get_db
from app.models.maintenance import (
    MaintenanceBudget,
    MaintenancePayment,
    MaintenanceServiceFee,
    RentalScheduleEntry,
)

router = APIRouter()

# ── Reference prefix ────────────────────────────────────────────────────────

async def _next_ref(db: AsyncSession, model, col_name: str, prefix: str) -> str:
    col = getattr(model, col_name)
    result = await db.execute(select(func.count()).select_from(model))
    count = (result.scalar() or 0) + 1
    return f"{prefix}-{count:04d}"


# ══════════════════════════════════════════════════════════════════════════════
# PAYMENTS
# ══════════════════════════════════════════════════════════════════════════════

class PaymentCreate(BaseModel):
    payment_reference: Optional[str] = None
    payment_type: str
    project: str
    block: Optional[str] = None
    unit: str
    payer_name: str
    payer_contact: Optional[str] = None
    amount: float
    currency: str = "GHS"
    payment_date: str
    payment_method: Optional[str] = None
    bank_name: Optional[str] = None
    cheque_number: Optional[str] = None
    received_by: Optional[str] = None
    description: Optional[str] = None
    notes: Optional[str] = None
    status: str = "completed"


@router.get("/payments")
async def list_payments(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    payment_type: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    q = select(MaintenancePayment).order_by(MaintenancePayment.created_at.desc())
    if payment_type:
        q = q.where(MaintenancePayment.payment_type == payment_type)
    if search:
        like = f"%{search}%"
        q = q.where(
            MaintenancePayment.payer_name.ilike(like)
            | MaintenancePayment.project.ilike(like)
            | MaintenancePayment.payment_reference.ilike(like)
            | MaintenancePayment.unit.ilike(like)
        )

    total_result = await db.execute(select(func.count()).select_from(q.subquery()))
    total = total_result.scalar() or 0

    q = q.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(q)
    items = result.scalars().all()

    def _row(p: MaintenancePayment):
        return {
            "id": str(p.id),
            "payment_reference": p.payment_reference,
            "payment_type": p.payment_type,
            "project": p.project,
            "block": p.block,
            "unit": p.unit,
            "payer_name": p.payer_name,
            "payer_contact": p.payer_contact,
            "amount": float(p.amount),
            "currency": p.currency,
            "payment_date": p.payment_date.isoformat() if p.payment_date else None,
            "payment_method": p.payment_method,
            "bank_name": p.bank_name,
            "cheque_number": p.cheque_number,
            "received_by": p.received_by,
            "description": p.description,
            "notes": p.notes,
            "status": p.status,
            "created_at": p.created_at.isoformat(),
        }

    return {"items": [_row(p) for p in items], "total": total, "page": page, "pages": max(1, -(-total // page_size))}


@router.post("/payments", status_code=status.HTTP_201_CREATED)
async def create_payment(
    data: PaymentCreate,
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    ref = data.payment_reference or await _next_ref(db, MaintenancePayment, "payment_reference", "MNT-PAY")
    try:
        pay_date = date.fromisoformat(data.payment_date)
    except (ValueError, TypeError):
        pay_date = date.today()

    payment = MaintenancePayment(
        payment_reference=ref,
        payment_type=data.payment_type,
        project=data.project,
        block=data.block,
        unit=data.unit,
        payer_name=data.payer_name,
        payer_contact=data.payer_contact,
        amount=Decimal(str(data.amount)),
        currency=data.currency,
        payment_date=pay_date,
        payment_method=data.payment_method,
        bank_name=data.bank_name,
        cheque_number=data.cheque_number,
        received_by=data.received_by,
        description=data.description,
        notes=data.notes,
        status=data.status,
        created_by=current_user,
    )
    db.add(payment)
    await db.commit()
    await db.refresh(payment)
    return {"id": str(payment.id), "payment_reference": payment.payment_reference}


class PaymentUpdate(BaseModel):
    payment_type: Optional[str] = None
    project: Optional[str] = None
    block: Optional[str] = None
    unit: Optional[str] = None
    payer_name: Optional[str] = None
    payer_contact: Optional[str] = None
    amount: Optional[float] = None
    currency: Optional[str] = None
    payment_date: Optional[str] = None
    payment_method: Optional[str] = None
    bank_name: Optional[str] = None
    cheque_number: Optional[str] = None
    received_by: Optional[str] = None
    description: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None


@router.patch("/payments/{payment_id}")
async def update_payment(
    payment_id: UUID,
    data: PaymentUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    result = await db.execute(select(MaintenancePayment).where(MaintenancePayment.id == payment_id))
    payment = result.scalar_one_or_none()
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")

    for field, val in data.model_dump(exclude_none=True).items():
        if field == "amount":
            payment.amount = Decimal(str(val))
        elif field == "payment_date":
            try:
                payment.payment_date = date.fromisoformat(val)
            except (ValueError, TypeError):
                pass
        else:
            setattr(payment, field, val)
    payment.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(payment)

    return {
        "id": str(payment.id),
        "payment_reference": payment.payment_reference,
        "payment_type": payment.payment_type,
        "project": payment.project,
        "block": payment.block,
        "unit": payment.unit,
        "payer_name": payment.payer_name,
        "payer_contact": payment.payer_contact,
        "amount": float(payment.amount),
        "currency": payment.currency,
        "payment_date": payment.payment_date.isoformat() if payment.payment_date else None,
        "payment_method": payment.payment_method,
        "bank_name": payment.bank_name,
        "cheque_number": payment.cheque_number,
        "received_by": payment.received_by,
        "description": payment.description,
        "notes": payment.notes,
        "status": payment.status,
    }


@router.delete("/payments/{payment_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_payment(
    payment_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    result = await db.execute(select(MaintenancePayment).where(MaintenancePayment.id == payment_id))
    payment = result.scalar_one_or_none()
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    await db.delete(payment)
    await db.commit()


# ══════════════════════════════════════════════════════════════════════════════
# BUDGETS
# ══════════════════════════════════════════════════════════════════════════════

class BudgetCreate(BaseModel):
    title: str
    project_id: Optional[str] = None
    project_name: Optional[str] = None
    category: Optional[str] = None
    total_amount: float
    currency: str = "GHS"
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    notes: Optional[str] = None
    created_by_name: Optional[str] = None


class BudgetUpdate(BaseModel):
    title: Optional[str] = None
    project_name: Optional[str] = None
    category: Optional[str] = None
    total_amount: Optional[float] = None
    paid_amount: Optional[float] = None
    currency: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None


def _budget_row(b: MaintenanceBudget) -> dict:
    outstanding = float(b.total_amount or 0) - float(b.paid_amount or 0)
    return {
        "id": str(b.id),
        "budget_reference": b.budget_reference,
        "title": b.title,
        "project_id": str(b.project_id) if b.project_id else None,
        "project_name": b.project_name,
        "category": b.category,
        "total_amount": float(b.total_amount or 0),
        "paid_amount": float(b.paid_amount or 0),
        "outstanding": outstanding,
        "currency": b.currency,
        "start_date": b.start_date.isoformat() if b.start_date else None,
        "end_date": b.end_date.isoformat() if b.end_date else None,
        "status": b.status,
        "notes": b.notes,
        "created_by": str(b.created_by) if b.created_by else None,
        "created_by_name": b.created_by_name,
        "created_at": b.created_at.isoformat(),
    }


@router.get("/budgets")
async def list_budgets(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    status_filter: Optional[str] = Query(None, alias="status"),
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    q = select(MaintenanceBudget).order_by(MaintenanceBudget.created_at.desc())
    if status_filter:
        q = q.where(MaintenanceBudget.status == status_filter)
    if search:
        like = f"%{search}%"
        q = q.where(
            MaintenanceBudget.title.ilike(like)
            | MaintenanceBudget.project_name.ilike(like)
            | MaintenanceBudget.budget_reference.ilike(like)
        )

    total_result = await db.execute(select(func.count()).select_from(q.subquery()))
    total = total_result.scalar() or 0

    # Summary totals (all records, ignoring pagination)
    totals_result = await db.execute(
        select(
            func.sum(MaintenanceBudget.total_amount),
            func.sum(MaintenanceBudget.paid_amount),
        )
    )
    t_total, t_paid = totals_result.one()

    q = q.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(q)
    items = result.scalars().all()

    return {
        "items": [_budget_row(b) for b in items],
        "total": total,
        "page": page,
        "pages": max(1, -(-total // page_size)),
        "summary": {
            "total_budget": float(t_total or 0),
            "total_paid": float(t_paid or 0),
            "total_outstanding": float((t_total or 0) - (t_paid or 0)),
        },
    }


@router.post("/budgets", status_code=status.HTTP_201_CREATED)
async def create_budget(
    data: BudgetCreate,
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    ref = await _next_ref(db, MaintenanceBudget, "budget_reference", "MNT-BGT")
    proj_id = None
    if data.project_id:
        try:
            proj_id = UUID(data.project_id)
        except ValueError:
            pass

    budget = MaintenanceBudget(
        budget_reference=ref,
        title=data.title,
        project_id=proj_id,
        project_name=data.project_name,
        category=data.category,
        total_amount=Decimal(str(data.total_amount)),
        paid_amount=Decimal("0"),
        currency=data.currency,
        start_date=date.fromisoformat(data.start_date) if data.start_date else None,
        end_date=date.fromisoformat(data.end_date) if data.end_date else None,
        status="pending_approval",
        notes=data.notes,
        created_by=current_user,
        created_by_name=data.created_by_name,
    )
    db.add(budget)
    await db.commit()
    await db.refresh(budget)
    return _budget_row(budget)


@router.patch("/budgets/{budget_id}")
async def update_budget(
    budget_id: UUID,
    data: BudgetUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    result = await db.execute(select(MaintenanceBudget).where(MaintenanceBudget.id == budget_id))
    budget = result.scalar_one_or_none()
    if not budget:
        raise HTTPException(status_code=404, detail="Budget not found")

    if data.title is not None:
        budget.title = data.title
    if data.project_name is not None:
        budget.project_name = data.project_name
    if data.category is not None:
        budget.category = data.category
    if data.total_amount is not None:
        budget.total_amount = Decimal(str(data.total_amount))
    if data.paid_amount is not None:
        budget.paid_amount = Decimal(str(data.paid_amount))
    if data.currency is not None:
        budget.currency = data.currency
    if data.start_date is not None:
        budget.start_date = date.fromisoformat(data.start_date)
    if data.end_date is not None:
        budget.end_date = date.fromisoformat(data.end_date)
    if data.status is not None:
        budget.status = data.status
    if data.notes is not None:
        budget.notes = data.notes
    budget.updated_at = datetime.utcnow()

    await db.commit()
    await db.refresh(budget)
    return _budget_row(budget)


@router.delete("/budgets/{budget_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_budget(
    budget_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    result = await db.execute(select(MaintenanceBudget).where(MaintenanceBudget.id == budget_id))
    budget = result.scalar_one_or_none()
    if not budget:
        raise HTTPException(status_code=404, detail="Budget not found")
    await db.delete(budget)
    await db.commit()


# ══════════════════════════════════════════════════════════════════════════════
# SERVICE FEES
# ══════════════════════════════════════════════════════════════════════════════

class ServiceFeeCreate(BaseModel):
    project: str
    block: Optional[str] = None
    unit: str
    owner_name: str
    owner_contact: Optional[str] = None
    fee_type: Optional[str] = None
    amount: float
    currency: str = "GHS"
    billing_period: Optional[str] = None
    due_date: Optional[str] = None
    status: str = "pending"
    payment_date: Optional[str] = None
    receipt_number: Optional[str] = None
    notes: Optional[str] = None


def _fee_row(f: MaintenanceServiceFee) -> dict:
    return {
        "id": str(f.id),
        "project": f.project,
        "block": f.block,
        "unit": f.unit,
        "owner_name": f.owner_name,
        "owner_contact": f.owner_contact,
        "fee_type": f.fee_type,
        "amount": float(f.amount),
        "currency": f.currency,
        "billing_period": f.billing_period,
        "due_date": f.due_date.isoformat() if f.due_date else None,
        "status": f.status,
        "payment_date": f.payment_date.isoformat() if f.payment_date else None,
        "receipt_number": f.receipt_number,
        "notes": f.notes,
        "created_at": f.created_at.isoformat(),
    }


@router.get("/service-fees")
async def list_service_fees(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status_filter: Optional[str] = Query(None, alias="status"),
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    q = select(MaintenanceServiceFee).order_by(MaintenanceServiceFee.created_at.desc())
    if status_filter:
        q = q.where(MaintenanceServiceFee.status == status_filter)
    if search:
        like = f"%{search}%"
        q = q.where(
            MaintenanceServiceFee.owner_name.ilike(like)
            | MaintenanceServiceFee.project.ilike(like)
            | MaintenanceServiceFee.unit.ilike(like)
        )

    total_result = await db.execute(select(func.count()).select_from(q.subquery()))
    total = total_result.scalar() or 0

    q = q.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(q)
    items = result.scalars().all()

    return {"items": [_fee_row(f) for f in items], "total": total, "page": page, "pages": max(1, -(-total // page_size))}


@router.post("/service-fees", status_code=status.HTTP_201_CREATED)
async def create_service_fee(
    data: ServiceFeeCreate,
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    fee = MaintenanceServiceFee(
        project=data.project,
        block=data.block,
        unit=data.unit,
        owner_name=data.owner_name,
        owner_contact=data.owner_contact,
        fee_type=data.fee_type,
        amount=Decimal(str(data.amount)),
        currency=data.currency,
        billing_period=data.billing_period,
        due_date=date.fromisoformat(data.due_date) if data.due_date else None,
        status=data.status,
        payment_date=date.fromisoformat(data.payment_date) if data.payment_date else None,
        receipt_number=data.receipt_number,
        notes=data.notes,
        created_by=current_user,
    )
    db.add(fee)
    await db.commit()
    await db.refresh(fee)
    return _fee_row(fee)


class ServiceFeeUpdate(BaseModel):
    project: Optional[str] = None
    block: Optional[str] = None
    unit: Optional[str] = None
    owner_name: Optional[str] = None
    owner_contact: Optional[str] = None
    fee_type: Optional[str] = None
    amount: Optional[float] = None
    currency: Optional[str] = None
    billing_period: Optional[str] = None
    due_date: Optional[str] = None
    status: Optional[str] = None
    payment_date: Optional[str] = None
    receipt_number: Optional[str] = None
    notes: Optional[str] = None


@router.patch("/service-fees/{fee_id}")
async def update_service_fee(
    fee_id: UUID,
    data: ServiceFeeUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    result = await db.execute(select(MaintenanceServiceFee).where(MaintenanceServiceFee.id == fee_id))
    fee = result.scalar_one_or_none()
    if not fee:
        raise HTTPException(status_code=404, detail="Service fee not found")

    for field, val in data.model_dump(exclude_none=True).items():
        if field == "amount":
            fee.amount = Decimal(str(val))
        elif field in ("due_date", "payment_date"):
            try:
                setattr(fee, field, date.fromisoformat(val))
            except (ValueError, TypeError):
                pass
        else:
            setattr(fee, field, val)
    fee.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(fee)
    return _fee_row(fee)


@router.delete("/service-fees/{fee_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_service_fee(
    fee_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    result = await db.execute(select(MaintenanceServiceFee).where(MaintenanceServiceFee.id == fee_id))
    fee = result.scalar_one_or_none()
    if not fee:
        raise HTTPException(status_code=404, detail="Service fee not found")
    await db.delete(fee)
    await db.commit()


# ══════════════════════════════════════════════════════════════════════════════
# SERVICE FEE IMPORT (Excel)
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/import/service-fees")
async def import_service_fees(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    """Import service fee records from an Excel file.
    Expected columns (flexible matching): Project, Block, Unit, Owner Name,
    Owner Contact, Fee Type, Amount, Currency, Billing Period, Due Date,
    Status, Payment Date, Receipt Number, Notes.
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx or .xls file")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed on server")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse Excel file: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="File is empty")

    # Find header row
    header_idx = None
    for i, row in enumerate(rows):
        row_text = " ".join(str(c or "").lower() for c in row)
        if ("unit" in row_text or "owner" in row_text) and ("amount" in row_text or "fee" in row_text):
            header_idx = i
            break

    if header_idx is None:
        header_idx = 0

    headers = [str(c or "").strip().lower() for c in rows[header_idx]]

    def col(frags: list[str]) -> Optional[int]:
        for i, h in enumerate(headers):
            if any(f in h for f in frags):
                return i
        return None

    c_project = col(["project", "property"])
    c_block = col(["block", "building"])
    c_unit = col(["unit", "apartment", "suite"])
    c_owner = col(["owner name", "owner"])
    c_contact = col(["contact", "phone", "email"])
    c_fee_type = col(["fee type", "type"])
    c_amount = col(["amount"])
    c_currency = col(["currency"])
    c_period = col(["billing period", "period", "billing"])
    c_due = col(["due date", "due"])
    c_status = col(["status"])
    c_pay_date = col(["payment date", "paid date"])
    c_receipt = col(["receipt", "reference"])
    c_notes = col(["notes", "remarks"])

    count = 0
    for row in rows[header_idx + 1:]:
        if not any(row):
            continue
        raw_amount = row[c_amount] if c_amount is not None else None
        amt = _parse_decimal(raw_amount)
        if amt is None:
            continue

        fee = MaintenanceServiceFee(
            project=str(row[c_project] or "").strip() if c_project is not None else "",
            block=str(row[c_block] or "").strip() if c_block is not None else None,
            unit=str(row[c_unit] or "").strip() if c_unit is not None else "",
            owner_name=str(row[c_owner] or "").strip() if c_owner is not None else "",
            owner_contact=str(row[c_contact] or "").strip() if c_contact is not None else None,
            fee_type=str(row[c_fee_type] or "").strip() if c_fee_type is not None else None,
            amount=amt,
            currency=str(row[c_currency] or "GHS").strip() if c_currency is not None else "GHS",
            billing_period=str(row[c_period] or "").strip() if c_period is not None else None,
            due_date=_parse_date(row[c_due]) if c_due is not None else None,
            status=str(row[c_status] or "pending").strip().lower() if c_status is not None else "pending",
            payment_date=_parse_date(row[c_pay_date]) if c_pay_date is not None else None,
            receipt_number=str(row[c_receipt] or "").strip() if c_receipt is not None else None,
            notes=str(row[c_notes] or "").strip() if c_notes is not None else None,
            created_by=current_user,
        )
        db.add(fee)
        count += 1

    await db.commit()
    return {"rows_imported": count, "sheets_processed": [ws.title]}


# ══════════════════════════════════════════════════════════════════════════════
# RENTAL SCHEDULE IMPORT (Excel)
# ══════════════════════════════════════════════════════════════════════════════

# Keyword-based property detection — sheet names don't need to match exactly.
# Each entry: (canonical_property_name, list_of_lowercase_keywords)
_PROPERTY_KEYWORDS = [
    ("Imperial Court", ["imperial court"]),
    ("Palazzo", ["palazzo"]),
    ("Philippa", ["philippa"]),
]


def _detect_property(sheet_name: str) -> Optional[tuple[str, str]]:
    """Return (property_name, year) if sheet_name contains a known property keyword."""
    lower = sheet_name.lower()
    for prop_name, keywords in _PROPERTY_KEYWORDS:
        if any(k in lower for k in keywords):
            year_match = re.search(r"\b(20\d{2})\b", sheet_name)
            year = year_match.group(1) if year_match else str(datetime.now().year)
            return (prop_name, year)
    return None


def _parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, (datetime,)):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(val) -> Optional[Decimal]:
    if val is None:
        return None
    try:
        return Decimal(str(val).replace(",", "").strip())
    except InvalidOperation:
        return None


@router.post("/import/rental-schedule")
async def import_rental_schedule(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    """Import a Palazzo-style rental schedule Excel file.
    Processes sheets: Palazzo 2026, Imperial Court 2026, Philippa 2026.
    Each sheet must have columns: Commercial/Unit, Sq M, Owner, Tenant,
    Start Date, Expiry Date, Months, Monthly Rent, Total, Amount Paid, Balance,
    Tenancy Agreement Status, Due Date, Status.
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx or .xls file")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed on server")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse Excel file: {e}")

    imported = []
    skipped = []

    for sheet_name in wb.sheetnames:
        detected = _detect_property(sheet_name)
        if not detected:
            skipped.append(sheet_name)
            continue

        property_name, sheet_year = detected

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find the header row — look for a row containing "Tenant" and date markers
        header_idx = None
        for i, row in enumerate(rows):
            row_text = " ".join(str(c or "").lower() for c in row)
            if "tenant" in row_text and ("start" in row_text or "expiry" in row_text):
                header_idx = i
                break

        if header_idx is None:
            skipped.append(f"{sheet_name} (header not found)")
            continue

        headers = [str(c or "").strip().lower() for c in rows[header_idx]]

        def col(name_fragments: list[str]) -> Optional[int]:
            for i, h in enumerate(headers):
                if any(f in h for f in name_fragments):
                    return i
            return None

        c_unit = col(["commercial", "suite", "unit", "block"])
        c_sqm = col(["square", "sqm", "sq m", "sq."])
        c_owner = col(["owner"])
        c_tenant = col(["tenant"])
        c_start = col(["start"])
        c_expiry = col(["expiry", "end"])
        c_months = col(["month"])
        c_rent = col(["monthly", "rent"])
        c_total = col(["total"])
        c_paid = col(["paid", "amount paid"])
        c_balance = col(["balance"])
        c_ta_status = col(["tenancy", "agreement", "status on"])
        c_due = col(["due"])
        c_status = col(["status"])

        count = 0
        for row in rows[header_idx + 1 :]:
            # Skip empty or total rows
            if not any(row):
                continue
            if c_unit is not None and str(row[c_unit] or "").strip().upper() in ("", "TOTAL", "COMMERCIAL"):
                continue

            entry = RentalScheduleEntry(
                property_name=property_name,
                sheet_year=sheet_year,
                commercial_unit=str(row[c_unit] or "").strip() if c_unit is not None else None,
                square_meters=str(row[c_sqm] or "").strip() if c_sqm is not None else None,
                owner=str(row[c_owner] or "").strip() if c_owner is not None else None,
                tenant=str(row[c_tenant] or "").strip() if c_tenant is not None else None,
                start_date=_parse_date(row[c_start]) if c_start is not None else None,
                expiry_date=_parse_date(row[c_expiry]) if c_expiry is not None else None,
                months=_parse_decimal(row[c_months]) if c_months is not None else None,
                monthly_rent=_parse_decimal(row[c_rent]) if c_rent is not None else None,
                currency="USD",
                total_amount=_parse_decimal(row[c_total]) if c_total is not None else None,
                amount_paid=_parse_decimal(row[c_paid]) if c_paid is not None else None,
                balance=_parse_decimal(row[c_balance]) if c_balance is not None else None,
                tenancy_agreement_status=str(row[c_ta_status] or "").strip() if c_ta_status is not None else None,
                due_date=_parse_date(row[c_due]) if c_due is not None else None,
                status_notes=str(row[c_status] or "").strip() if c_status is not None else None,
                imported_by=current_user,
            )
            db.add(entry)
            count += 1

        imported.append({"sheet": sheet_name, "rows_imported": count})

    await db.commit()
    return {"imported": imported, "skipped": skipped}


@router.get("/rental-schedule")
async def list_rental_schedule(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    property_name: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: UUID = Depends(get_current_user),
):
    q = select(RentalScheduleEntry).where(RentalScheduleEntry.is_active == True).order_by(
        RentalScheduleEntry.property_name, RentalScheduleEntry.commercial_unit
    )
    if property_name:
        q = q.where(RentalScheduleEntry.property_name.ilike(f"%{property_name}%"))
    if search:
        like = f"%{search}%"
        q = q.where(
            RentalScheduleEntry.tenant.ilike(like)
            | RentalScheduleEntry.commercial_unit.ilike(like)
            | RentalScheduleEntry.owner.ilike(like)
        )

    total_result = await db.execute(select(func.count()).select_from(q.subquery()))
    total = total_result.scalar() or 0

    q = q.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(q)
    items = result.scalars().all()

    def _row(e: RentalScheduleEntry) -> dict:
        return {
            "id": str(e.id),
            "property_name": e.property_name,
            "sheet_year": e.sheet_year,
            "commercial_unit": e.commercial_unit,
            "square_meters": e.square_meters,
            "owner": e.owner,
            "tenant": e.tenant,
            "start_date": e.start_date.isoformat() if e.start_date else None,
            "expiry_date": e.expiry_date.isoformat() if e.expiry_date else None,
            "months": float(e.months) if e.months else None,
            "monthly_rent": float(e.monthly_rent) if e.monthly_rent else None,
            "currency": e.currency,
            "total_amount": float(e.total_amount) if e.total_amount else None,
            "amount_paid": float(e.amount_paid) if e.amount_paid else None,
            "balance": float(e.balance) if e.balance else None,
            "tenancy_agreement_status": e.tenancy_agreement_status,
            "due_date": e.due_date.isoformat() if e.due_date else None,
            "status_notes": e.status_notes,
        }

    return {"items": [_row(e) for e in items], "total": total, "page": page, "pages": max(1, -(-total // page_size))}
